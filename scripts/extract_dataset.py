#!/usr/bin/env python3
"""Extract the DGII test dataset (.xlsx) into clean per-case JSON.

The .xlsx has two sheets, ECF and RFCE. Each row is one test case. Columns are
the flattened e-CF field names. We emit a JSON list per sheet where each case is
{ fieldName: value } including ONLY cells that actually have a value (so the
generator can omit empty/`#e` fields and include every populated one — which is
exactly what fixes the 'missing_tag' rejections).
"""
import openpyxl, json, sys, os

SRC = sys.argv[1] if len(sys.argv) > 1 else "dataset.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "dataset.json"

def clean(v):
    if v is None:
        return None
    if isinstance(v, float):
        # keep integers clean, money to 2 decimals handled later by generator
        if v.is_integer():
            return str(int(v))
        return repr(v)
    s = str(v).strip()
    return s if s != "" else None

wb = openpyxl.load_workbook(SRC, data_only=True)
out = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    hdr = [clean(h) for h in rows[0]]
    cases = []
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue
        case = {}
        for i, val in enumerate(r):
            if i >= len(hdr):
                break
            key = hdr[i]
            if not key:
                continue
            cv = clean(val)
            if cv is not None:
                # strip stray trailing spaces in header names too (NumeroContenedor )
                case[key.strip()] = cv
        cases.append(case)
    out[sheet] = cases
    print(f"{sheet}: {len(cases)} cases, {len([h for h in hdr if h])} named columns", file=sys.stderr)

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"wrote {OUT}", file=sys.stderr)
